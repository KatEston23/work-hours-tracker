import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar
from typing import List, Tuple, Dict, Optional


class WorkHoursTracker:
    """Manages work hours tracking and Excel calendar generation."""
    
    EXCEL_FILENAME = "work_hours_history.xlsx"
    STANDARD_WORKDAY_HOURS = 8
    
    def __init__(self):
        self.selected_date = None
        self.time_entries = []
        
    def run(self):
        """Main application workflow."""
        self.selected_date = self._get_date_from_user()
        self.time_entries = self._collect_time_entries()
        
        if not self.time_entries:
            print("⚠️ No time entries recorded.")
            return
        
        worked_hours, extra_hours = self._calculate_hours()
        self._display_results(worked_hours, extra_hours)
        self._save_to_excel(worked_hours, extra_hours)
    
    def _get_date_from_user(self) -> str:
        """Get date selection from user input."""
        print("📅 DATE SELECTION")
        print("1. Use today's date")
        print("2. Yesterday")
        print("3. Enter custom date")
        
        choice = input("Choose option (1, 2, or 3): ").strip()
        
        if choice == "1":
            return self._get_today_date()
        elif choice == "2":
            return self._get_yesterday_date()
        elif choice == "3":
            return self._get_custom_date()
        else:
            print("⚠ Invalid choice. Using today's date as default.")
            return self._get_today_date()
    
    def _get_today_date(self) -> str:
        """Get today's date."""
        return datetime.today().strftime("%Y-%m-%d")
    
    def _get_yesterday_date(self) -> str:
        """Get yesterday's date."""
        yesterday = datetime.today() - timedelta(days=1)
        return yesterday.strftime("%Y-%m-%d")
    
    def _get_custom_date(self) -> str:
        """Get and validate custom date from user."""
        while True:
            date_input = input("Enter date (YYYY-MM-DD format, example: 2025-01-15): ").strip()
            
            if self._is_valid_date(date_input):
                return date_input
            
            print("⚠️ Invalid date format. Please use YYYY-MM-DD (example: 2025-01-15)")
    
    def _is_valid_date(self, date_string: str) -> bool:
        """Validate date format and ensure it's not in the future."""
        try:
            date_object = datetime.strptime(date_string, "%Y-%m-%d")
            
            if date_object > datetime.today():
                print("⚠ Cannot enter hours for future dates. Please select today or a past date.")
                return False
                
            return True
        except ValueError:
            return False
    
    def _collect_time_entries(self) -> List[List[str]]:
        """Collect time entries from user input."""
        date_object = datetime.strptime(self.selected_date, "%Y-%m-%d")
        day_name = date_object.strftime("%A")
        
        print(f"\n📅 Selected date: {self.selected_date} ({day_name})")
        print("Enter your time entries in HH.MM format (example: 08.42).")
        print("Type 'done' when finished.\n")
        
        entries = []
        entry_counter = 1
        
        while True:
            time_input = input(f"Time entry {entry_counter}: ")
            
            if time_input.lower() == "done":
                break
                
            if self._is_valid_time_format(time_input):
                entry_type = self._determine_entry_type(len(entries))
                entries.append([self.selected_date, entry_type, time_input])
                entry_counter += 1
            else:
                print("⚠️ Invalid format. Use HH.MM (example: 07.30)")
        
        return entries
    
    def _is_valid_time_format(self, time_string: str) -> bool:
        """Validate time format HH.MM."""
        try:
            datetime.strptime(time_string, "%H.%M")
            return True
        except ValueError:
            return False
    
    def _determine_entry_type(self, entry_count: int) -> str:
        """Determine if entry is clock in or clock out."""
        return "Clock In" if entry_count % 2 == 0 else "Clock Out"
    
    def _calculate_hours(self) -> Tuple[timedelta, timedelta]:
        """Calculate total worked hours and extra hours."""
        time_dataframe = self._create_time_dataframe()
        total_worked = self._calculate_total_worked_time(time_dataframe)
        extra_hours = total_worked - timedelta(hours=self.STANDARD_WORKDAY_HOURS)
        
        return total_worked, extra_hours
    
    def _create_time_dataframe(self) -> pd.DataFrame:
        """Create DataFrame from time entries."""
        dataframe = pd.DataFrame(self.time_entries, columns=["Date", "Type", "Time"])
        dataframe["DateTime"] = pd.to_datetime(
            dataframe["Date"] + " " + dataframe["Time"], 
            format="%Y-%m-%d %H.%M"
        )
        return dataframe
    
    def _calculate_total_worked_time(self, time_dataframe: pd.DataFrame) -> timedelta:
        """Calculate total worked time from clock in/out pairs."""
        total_worked = timedelta()
        
        for i in range(0, len(time_dataframe), 2):
            if i + 1 < len(time_dataframe):
                clock_in_time = time_dataframe.loc[i, "DateTime"]
                clock_out_time = time_dataframe.loc[i + 1, "DateTime"]
                total_worked += clock_out_time - clock_in_time
        
        return total_worked
    
    def _display_results(self, worked_hours: timedelta, extra_hours: timedelta):
        """Display calculation results to user."""
        print("\n📊 RESULTS")
        print(f"Hours worked: {self._format_timedelta(worked_hours)}")
        print(f"Extra hours: {self._format_timedelta(extra_hours)}")
    
    def _format_timedelta(self, time_delta: timedelta) -> str:
        """Convert timedelta to readable HH:MM format."""
        total_seconds = int(time_delta.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    
    def _save_to_excel(self, worked_hours: timedelta, extra_hours: timedelta):
        """Save work hours data to Excel calendar format."""
        worked_hours_string = self._format_timedelta(worked_hours)
        extra_hours_string = self._format_timedelta(extra_hours)
        
        excel_generator = ExcelCalendarGenerator()
        excel_generator.create_calendar_excel(
            self.EXCEL_FILENAME, 
            self.selected_date, 
            worked_hours_string, 
            extra_hours_string
        )
        
        print(f"\n✅ Data saved to '{self.EXCEL_FILENAME}' in calendar format.")


class ExcelCalendarGenerator:
    """Handles Excel calendar creation and formatting."""
    
    def create_calendar_excel(self, file_path: str, date: str, worked_hours: str, extra_hours: str):
        """Create Excel file with calendar format."""
        work_data = self._load_existing_data(file_path)
        work_data = self._add_current_day_data(work_data, date, worked_hours, extra_hours)
        
        workbook = self._create_workbook_with_calendars(work_data)
        self._add_backup_data_sheet(workbook, work_data)
        workbook.save(file_path)
    
    def _load_existing_data(self, file_path: str) -> Dict[str, Dict[str, Dict[str, str]]]:
        """Load existing work data from Excel file."""
        if not os.path.exists(file_path):
            return {}
        
        try:
            return self._load_from_data_sheet(file_path)
        except:
            try:
                return self._load_from_legacy_format(file_path)
            except:
                return {}
    
    def _load_from_data_sheet(self, file_path: str) -> Dict[str, Dict[str, Dict[str, str]]]:
        """Load data from the Data sheet."""
        existing_dataframe = pd.read_excel(file_path, sheet_name='Data')
        return self._convert_dataframe_to_work_data(existing_dataframe)
    
    def _load_from_legacy_format(self, file_path: str) -> Dict[str, Dict[str, Dict[str, str]]]:
        """Load data from legacy Excel format."""
        old_dataframe = pd.read_excel(file_path)
        return self._convert_dataframe_to_work_data(old_dataframe)
    
    def _convert_dataframe_to_work_data(self, dataframe: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, str]]]:
        """Convert DataFrame to work data dictionary structure."""
        work_data = {}
        
        for _, row in dataframe.iterrows():
            date_object = self._parse_date_from_row(row['Date'])
            month_year = date_object.strftime("%Y-%m")
            date_string = date_object.strftime("%Y-%m-%d")
            
            if month_year not in work_data:
                work_data[month_year] = {}
            
            work_data[month_year][date_string] = {
                'worked_hours': row['Hours Worked'],
                'extra_hours': row['Extra Hours']
            }
        
        return work_data
    
    def _parse_date_from_row(self, date_value) -> datetime:
        """Parse date from Excel row, handling both string and datetime types."""
        if isinstance(date_value, str):
            return datetime.strptime(date_value, "%Y-%m-%d")
        return date_value
    
    def _add_current_day_data(self, work_data: dict, date: str, worked_hours: str, extra_hours: str) -> dict:
        """Add today's work data to the existing data."""
        date_object = datetime.strptime(date, "%Y-%m-%d")
        month_year = date_object.strftime("%Y-%m")
        
        if month_year not in work_data:
            work_data[month_year] = {}
        
        work_data[month_year][date] = {
            'worked_hours': worked_hours,
            'extra_hours': extra_hours
        }
        
        return work_data
    
    def _create_workbook_with_calendars(self, work_data: dict) -> Workbook:
        """Create workbook with calendar sheets for each month."""
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        for month_key in sorted(work_data.keys()):
            year, month = map(int, month_key.split('-'))
            sheet_name = f"{calendar.month_name[month]} {year}"
            worksheet = workbook.create_sheet(title=sheet_name)
            
            calendar_formatter = MonthlyCalendarFormatter()
            calendar_formatter.create_monthly_calendar(worksheet, year, month, work_data[month_key])
        
        return workbook
    
    def _add_backup_data_sheet(self, workbook: Workbook, work_data: dict):
        """Add backup data sheet with all records."""
        data_worksheet = workbook.create_sheet(title='Data')
        all_records = self._flatten_work_data(work_data)
        
        if all_records:
            backup_dataframe = pd.DataFrame(all_records, columns=['Date', 'Hours Worked', 'Extra Hours'])
            for row in dataframe_to_rows(backup_dataframe, index=False, header=True):
                data_worksheet.append(row)
    
    def _flatten_work_data(self, work_data: dict) -> List[List[str]]:
        """Flatten work data dictionary to list of records."""
        all_records = []
        for month_data in work_data.values():
            for date, hours in month_data.items():
                all_records.append([date, hours['worked_hours'], hours['extra_hours']])
        return all_records


class MonthlyCalendarFormatter:
    """Handles formatting of monthly calendar sheets."""
    
    def create_monthly_calendar(self, worksheet, year: int, month: int, month_data: dict):
        """Create formatted monthly calendar in worksheet."""
        self._add_month_title(worksheet, year, month)
        self._add_day_headers(worksheet)
        self._fill_calendar_days(worksheet, year, month, month_data)
        self._add_monthly_totals(worksheet, month_data, len(calendar.monthcalendar(year, month)))
        self._adjust_column_widths(worksheet)
    
    def _add_month_title(self, worksheet, year: int, month: int):
        """Add month title to worksheet."""
        month_name = f"{calendar.month_name[month]} {year}"
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = month_name
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def _add_day_headers(self, worksheet):
        """Add day of week headers."""
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = self._create_thin_border()
        
        for i, day_name in enumerate(day_names):
            cell = worksheet.cell(row=3, column=i+1, value=day_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    def _fill_calendar_days(self, worksheet, year: int, month: int, month_data: dict):
        """Fill calendar with days and work hours data."""
        calendar_weeks = calendar.monthcalendar(year, month)
        
        for week_number, week in enumerate(calendar_weeks):
            for day_number, day in enumerate(week):
                if day != 0:
                    self._add_day_cell(worksheet, week_number, day_number, day, year, month, month_data)
    
    def _add_day_cell(self, worksheet, week_number: int, day_number: int, day: int, year: int, month: int, month_data: dict):
        """Add individual day cell with work hours if available."""
        row = 4 + week_number
        column = day_number + 1
        date_string = f"{year}-{month:02d}-{day:02d}"
        
        worksheet.cell(row=row, column=column, value=day).font = Font(bold=True)
        
        if date_string in month_data:
            self._add_work_hours_to_cell(worksheet, row, column, month_data[date_string])
        
        self._apply_cell_formatting(worksheet, row, column, day_number >= 5)
    
    def _add_work_hours_to_cell(self, worksheet, row: int, column: int, day_data: dict):
        """Add work hours information to day cell."""
        worksheet.cell(row=row+1, column=column, value=f"W: {day_data['worked_hours']}")
        
        if day_data['extra_hours'] != "00:00":
            extra_cell = worksheet.cell(row=row+2, column=column, value=f"E: {day_data['extra_hours']}")
            extra_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            extra_cell.fill = extra_fill
    
    def _apply_cell_formatting(self, worksheet, row: int, column: int, is_weekend: bool):
        """Apply background color and border formatting to cells."""
        weekend_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        workday_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        thin_border = self._create_thin_border()
        
        fill_color = weekend_fill if is_weekend else workday_fill
        
        for cell_row in range(row, row + 3):
            cell = worksheet.cell(row=cell_row, column=column)
            cell.fill = fill_color
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    def _add_monthly_totals(self, worksheet, month_data: dict, calendar_weeks_count: int):
        """Add monthly totals row to worksheet."""
        total_row = 4 + calendar_weeks_count + 1
        total_worked, total_extra = self._calculate_monthly_totals(month_data)
        
        worksheet.merge_cells(f'A{total_row}:B{total_row}')
        worksheet[f'A{total_row}'] = "MONTHLY TOTAL:"
        worksheet[f'A{total_row}'].font = Font(bold=True)
        
        worked_cell = worksheet[f'C{total_row}']
        worked_cell.value = f"Worked: {self._format_timedelta(total_worked)}"
        worked_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        worked_cell.font = Font(bold=True)
        
        extra_cell = worksheet[f'E{total_row}']
        extra_cell.value = f"Extra: {self._format_timedelta(total_extra)}"
        extra_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        extra_cell.font = Font(bold=True)
    
    def _calculate_monthly_totals(self, month_data: dict) -> Tuple[timedelta, timedelta]:
        """Calculate total worked and extra hours for the month."""
        total_worked = timedelta()
        total_extra = timedelta()
        
        for day_data in month_data.values():
            total_worked += self._parse_time_string(day_data['worked_hours'])
            total_extra += self._parse_time_string(day_data['extra_hours'])
        
        return total_worked, total_extra
    
    def _adjust_column_widths(self, worksheet):
        """Adjust column widths for better readability."""
        for column_number in range(1, 8):
            worksheet.column_dimensions[chr(64 + column_number)].width = 15
    
    def _create_thin_border(self) -> Border:
        """Create thin border style."""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _parse_time_string(self, time_string: str) -> timedelta:
        """Convert time string HH:MM to timedelta."""
        try:
            hours, minutes = map(int, time_string.split(':'))
            return timedelta(hours=hours, minutes=minutes)
        except:
            return timedelta()
    
    def _format_timedelta(self, time_delta: timedelta) -> str:
        """Convert timedelta to readable HH:MM format."""
        total_seconds = int(time_delta.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"


if __name__ == "__main__":
    tracker = WorkHoursTracker()
    tracker.run()